# -*- coding: utf-8 -*-
"""สร้างไฟล์ manual test สำหรับ role Dean (คณบดี) — ฟอร์แมตเดียวกับชุดก่อน"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROWS = [
("▶ การเข้าสู่ระบบ (Dean = 41172017 / Test@1234)", None),
(1, "ล็อกอินด้วย 41172017 / Test@1234", "เข้าสู่ระบบได้ และถูกพาไปหน้า 'แดชบอร์ดคณบดี' อัตโนมัติ (role อาจารย์ + ตำแหน่งคณบดี)"),
(2, "ดูชื่อผู้ใช้มุมซ้ายล่างของ sidebar", "แสดงชื่อ 'อรทิพา ส่องศิริ'"),
(3, "ดูเมนู sidebar", "เห็นเมนูฝั่งบริหาร เช่น แดชบอร์ด KPI คณะ, อัตราคงอยู่ของ นศ. (เมนูมาจากสิทธิ์ของตำแหน่ง)"),
("▶ แดชบอร์ดคณบดี (?page=dean-dashboard)", None),
(4, "เปิดหน้าแดชบอร์ด", "เห็นหัวข้อ 'แดชบอร์ดคณบดี', ปีการศึกษา, ปุ่ม Export และการ์ดสถิติ 4 ใบ"),
(5, "ดูการ์ด 'นักศึกษาทั้งหมด'", "แสดง 131 คน (ข้อมูลจริงจากฐานข้อมูล)"),
(6, "ดูการ์ด 'อาจารย์ทั้งหมด'", "แสดง 53 คน (ข้อมูลชุด clean)"),
(7, "ดูการ์ด 'อัตราคงอยู่'", "แสดง 100% (นักศึกษาทุกคนสถานะกำลังศึกษา)"),
(8, "ดูการ์ด 'งบประมาณโครงการรวม'", "แสดง ฿350,000 (งบโครงการ A ปี 2566+2567)"),
(9, "กดแท็บ 'KPI Dashboard'", "แสดงกราฟ KPI นักศึกษา/อาจารย์ (หมายเหตุ: กราฟชุดนี้ยังเป็นข้อมูลตัวอย่างในตัว component)"),
(10, "กดแท็บ 'การคงอยู่'", "แสดงกราฟการคงอยู่และสาเหตุการพ้นสภาพ (กราฟยังเป็นข้อมูลตัวอย่าง)"),
(11, "กดแท็บ 'รายงานการเงิน'", "แสดงรายงานการเงินและสรุปผู้บริหาร (ยังเป็นข้อมูลตัวอย่าง)"),
(12, "กดปุ่ม Export มุมขวาบน", "สังเกตพฤติกรรม (ถ้าไม่มีไฟล์ดาวน์โหลด ให้บันทึกไว้ — ฟีเจอร์อาจยังไม่เสร็จ)"),
(13, "รีเฟรชหน้า (F5)", "โหลดข้อมูลใหม่ได้ ไม่ error ไม่หลุด login"),
("▶ อัตราคงอยู่ (?page=retention)", None),
(14, "เปิดหน้าอัตราคงอยู่จากเมนู", "หน้าเปิดได้ ไม่ error (เนื้อหาตามที่ทีมพัฒนาไว้)"),
("▶ สิทธิ์การเข้าถึง (ทดสอบด้วยบัญชีอื่น)", None),
(15, "ล็อกอิน admin 46172040 แล้วเปิด ?page=dean-dashboard", "Admin เปิดดูได้ (สิทธิ์ผู้ดูแลระบบเข้าได้ทุกหน้า)"),
(16, "ล็อกอินอาจารย์ทั่วไป 41172008 แล้วเปิด ?page=dean-dashboard", "ขึ้นหน้า 'ปฏิเสธการเข้าถึง (Access Denied)' เพราะไม่ใช่คณบดี"),
(17, "ล็อกอินนักศึกษา 6603400001 แล้วเปิด ?page=dean-dashboard", "ขึ้นหน้า 'ปฏิเสธการเข้าถึง' เช่นกัน"),
(18, "เปิด ?page=dean-dashboard โดยไม่ล็อกอิน (เปิดแท็บ incognito)", "ไม่แสดงข้อมูล KPI (ถูกพาไปหน้า login หรือแจ้งไม่ได้รับอนุญาต)"),
]

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Nursing_Dean_Manual_Test.xlsx")
wb = Workbook(); ws = wb.active; ws.title = "แบบทดสอบ Dean"
hdr_fill = PatternFill("solid", fgColor="8A2BE2"); hdr_font = Font(bold=True, color="FFFFFF")
mod_fill = PatternFill("solid", fgColor="EDE3FB"); mod_font = Font(bold=True, color="4B0082")
thin = Side(style="thin", color="D0D0D0"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
top = Alignment(vertical="top", wrap_text=True)

headers = ["ข้อ", "สิ่งที่ต้องทำ", "ผลที่ควรเห็น", "ผ่าน/ไม่ผ่าน", "หมายเหตุ"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

r = 2
for row in ROWS:
    if row[1] is None:
        ws.cell(row=r, column=1, value=row[0])
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = mod_fill; ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=1).font = mod_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    else:
        for c, val in enumerate([row[0], row[1], row[2], "", ""], 1):
            cell = ws.cell(row=r, column=c, value=val); cell.border = border; cell.alignment = top
    r += 1

for i, w in enumerate([6, 45, 55, 12, 30], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
wb.save(OUT)
print("WROTE:", OUT, "| items:", sum(1 for x in ROWS if x[1] is not None))
